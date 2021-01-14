import itertools
import os
import string

from openpyxl import load_workbook


def excel_data(id_min, id_max, row_range=35):
    filename = os.path.join(os.path.join(os.path.abspath(os.path.dirname(__file__)), '..', 'uploads'), 'excel.xlsx')
    wb = load_workbook(filename)
    ws = wb.active
    id_range = [f'id{idx}' for idx in range(id_min, id_max + 1)]
    cords = []

    for idx in id_range:
        for col in ws.iter_rows():
            for cell in col:
                if cell.value == idx:
                    cw_field = cell.coordinate
                    cords.append(cw_field)


    cords_letters = []
    for item in cords:
        for _ in item:
            if not _.isdigit():
                cord_letter = _
                cords_letters.append(cord_letter)

    final_data = []
    for letter in cords_letters:
        data = []
        for digit in range(2, row_range + 1):
            cell = f'{letter}{digit}'
            data.append(ws[cell].value)
        final_data.append(data)

    return final_data


def generate_excel(data):
    filename = os.path.join(os.path.join(os.path.abspath(os.path.dirname(__file__)), '..', 'uploads'), 'excel.xlsx')
    wb = load_workbook(filename)
    ws = wb.active
    columns = list(itertools.chain(string.ascii_uppercase,
                                   (''.join(pair) for pair in itertools.product(string.ascii_uppercase, repeat=2))))

    digits = len(data)
    col = 0
    for item in data:
        col += 1
        for col in range(1, digits+1):
            ws[f'{columns[col]}1'] = f'id{col - 1}'
            ws[f'{columns[col]}2'] = item.cw
            ws[f'{columns[col]}3'] = item.total_client_req
            ws[f'{columns[col]}4'] = item.total_client_req_approved
            ws[f'{columns[col]}5'] = item.total_open_issue_client_req
            ws[f'{columns[col]}6'] = item.total_sys_req
            ws[f'{columns[col]}7'] = item.total_sys_req_approved
            ws[f'{columns[col]}8'] = item.total_sys_req_implemented
            ws[f'{columns[col]}9'] = item.total_sw_req
            ws[f'{columns[col]}10'] = item.total_sw_req_approved
            ws[f'{columns[col]}11'] = item.total_sw_req_implemented
            ws[f'{columns[col]}12'] = item.ccm_more_50
            ws[f'{columns[col]}13'] = item.ccm_24_50
            ws[f'{columns[col]}14'] = item.ccm_12_24
            ws[f'{columns[col]}15'] = item.misra_high
            ws[f'{columns[col]}16'] = item.misra_mid
            ws[f'{columns[col]}17'] = item.misra_low
            ws[f'{columns[col]}18'] = item.branch_coverage
            ws[f'{columns[col]}19'] = item.line_coverage
            ws[f'{columns[col]}20'] = item.mc_dc_coverage
            ws[f'{columns[col]}21'] = item.total_sw_req_tests
            ws[f'{columns[col]}22'] = item.total_sw_req_passed
            ws[f'{columns[col]}23'] = item.total_sys_req_tests
            ws[f'{columns[col]}24'] = item.total_sys_req_tests_passed
            ws[f'{columns[col]}25'] = item.traceability_sys_client
            ws[f'{columns[col]}26'] = item.traceability_client_sys
            ws[f'{columns[col]}27'] = item.traceability_sys_sw
            ws[f'{columns[col]}28'] = item.traceability_sw_sys
            ws[f'{columns[col]}29'] = item.total_features
            ws[f'{columns[col]}30'] = item.total_bugs
            ws[f'{columns[col]}31'] = item.total_bugs_solved
            ws[f'{columns[col]}32'] = item.total_problems
            ws[f'{columns[col]}33'] = item.total_problems_solved
            ws[f'{columns[col]}34'] = item.total_change_requests
            ws[f'{columns[col]}35'] = item.change_request_not_reviewed
        wb.save(filename)