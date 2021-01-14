import itertools
import os
import string
from datetime import datetime

from flask import Blueprint, render_template, url_for, jsonify, flash, request, send_file
from flask_login import login_required
from openpyxl import load_workbook
from werkzeug.utils import redirect

from .. import db
from ..forms.workbook_form import WorkBookForm, ExcelForm

from ..models.user_models import User
from ..models.workbook_models import Workbook
from ..helpers import excel_data, generate_excel

bp_input = Blueprint("input", __name__, url_prefix='/metrics')


@bp_input.route('/add', methods=["GET", "POST"])
def metrics():
    form = WorkBookForm()

    if form.validate_on_submit():
        workbook = Workbook()
        form.populate_obj(workbook)
        db.session.add(workbook)
        db.session.commit()

        return redirect(url_for('main.home'))
    return render_template('data_add.html', form=form, title="Add new data")


@bp_input.route('/edit/<int:workbook_id>', methods=["GET", "POST"])
def edit(workbook_id):
    workbook = Workbook.get_metric_by_id(workbook_id)
    form = WorkBookForm(obj=workbook)

    if form.validate_on_submit():
        form.populate_obj(workbook)
        db.session.commit()
        flash(f'Saved, description: {workbook.cw}')
        return redirect(url_for('input.raw'))

    return render_template("data_add.html", form=form, title="Edit data", workbook_id=workbook_id)


@bp_input.route('/delete/<int:workbook_id>', methods=['GET', 'POST'])
def delete(workbook_id):
    workbook = Workbook.get_metric_by_id(workbook_id)

    if request.method == "POST":
        db.session.delete(workbook)
        db.session.commit()
        flash(f'Data has been removed: {workbook.cw}.')
        return redirect(url_for('input.raw'))

    else:
        flash('Please confirm deleting the bookmark.')

    return render_template("confirm_delete.html", workbook=workbook)


@bp_input.route('/add/excel', methods=['GET', 'POST'])
@login_required
def excel():
    form = ExcelForm()
    if form.validate_on_submit():
        form.excel.data.save(
            os.path.join(os.path.join(os.path.abspath(os.path.dirname(__file__)), '..', 'uploads'), 'excel.xlsx'))

        excel_values = excel_data(form.id_min.data, form.id_max.data)

        for item in excel_values:

            if item[0] is None:
                item[0] = f'{datetime.utcnow().strftime("%m/%d/%Y")}'

            workbook = Workbook(
                cw=item[0],
                total_client_req=item[1],
                total_client_req_approved=item[2],
                total_open_issue_client_req=item[3],
                total_sys_req=item[4],
                total_sys_req_approved=item[5],
                total_sys_req_implemented=item[6],
                total_sw_req=item[7],
                total_sw_req_approved=item[8],
                total_sw_req_implemented=item[9],
                ccm_more_50=item[10],
                ccm_24_50=item[11],
                ccm_12_24=item[12],
                misra_high=item[13],
                misra_mid=item[14],
                misra_low=item[15],
                branch_coverage=item[16],
                line_coverage=item[17],
                mc_dc_coverage=item[18],
                total_sw_req_tests=item[19],
                total_sw_req_passed=item[20],
                total_sys_req_tests=item[21],
                total_sys_req_tests_passed=item[22],
                traceability_sys_client=item[23],
                traceability_client_sys=item[24],
                traceability_sys_sw=item[25],
                traceability_sw_sys=item[26],
                total_features=item[27],
                total_bugs=item[28],
                total_bugs_solved=item[29],
                total_problems=item[30],
                total_problems_solved=item[31],
                total_change_requests=item[32],
                change_request_not_reviewed=item[33],
            )
            db.session.add(workbook)
            db.session.commit()
            flash(f'Data id: {item[0]} has been added', 'success')
        return redirect(url_for('input.raw'))

    return render_template('add_data_excel.html', form=form)


@bp_input.route('/raw')
def raw():
    return render_template('raw.html', raw_data=Workbook.all_of_workbooks(), users=User.all_of_users())


@bp_input.route('/charts')
def chart():
    return render_template('charts.html')


@bp_input.route('/google/charts')
def google_chart():
    return render_template('charts_google.html')


@bp_input.route('/data')
def data():
    cw = []
    total_client_req = []
    total_client_req_approved = []
    total_open_issue_client_req = []
    total_sys_req = []
    total_sys_req_approved = []
    total_sys_req_implemented = []
    total_sw_req = []
    total_sw_req_approved = []
    total_sw_req_implemented = []
    ccm_more_50 = []
    ccm_24_50 = []
    ccm_12_24 = []
    misra_high = []
    misra_mid = []
    misra_low = []
    branch_coverage = []
    line_coverage = []
    mc_dc_coverage = []
    total_sw_req_tests = []
    total_sw_req_passed = []
    total_sys_req_tests = []
    total_sys_req_tests_passed = []
    traceability_sys_client = []
    traceability_client_sys = []
    traceability_sys_sw = []
    traceability_sw_sys = []
    total_features = []
    total_bugs = []
    total_bugs_solved = []
    total_problems = []
    total_problems_solved = []
    total_change_requests = []
    change_request_not_reviewed = []

    data = Workbook.all_of_workbooks()
    for row in data[::-1]:
        cw.append(row.cw)
        total_client_req.append(row.total_client_req)
        total_change_requests.append(row.total_change_requests)
        change_request_not_reviewed.append(row.change_request_not_reviewed)
        traceability_sys_client.append(row.traceability_sys_client)
        traceability_sw_sys.append(row.traceability_sw_sys)
        total_features.append(row.total_features)
        total_bugs.append(row.total_bugs)
        total_bugs_solved.append(row.total_bugs_solved)
        total_problems.append(row.total_problems)
        total_problems_solved.append(row.total_problems_solved)
        traceability_client_sys.append(row.traceability_client_sys)
        traceability_sys_sw.append(row.traceability_sys_sw)
        total_sw_req_tests.append(row.total_sw_req_tests)
        total_sw_req_passed.append(row.total_sw_req_passed)
        total_sys_req_tests.append(row.total_sys_req_tests)
        total_sys_req_tests_passed.append(row.total_sys_req_tests_passed)
        branch_coverage.append(row.branch_coverage)
        line_coverage.append(row.line_coverage)
        mc_dc_coverage.append(row.mc_dc_coverage)
        total_client_req_approved.append(row.total_client_req_approved)
        total_open_issue_client_req.append(row.total_open_issue_client_req)
        total_sys_req.append(row.total_sys_req)
        total_sys_req_implemented.append(row.total_sys_req_implemented)
        total_sw_req.append(row.total_sw_req)
        total_sys_req_approved.append(row.total_sys_req_approved)
        total_sw_req_approved.append(row.total_sw_req_approved)
        total_sw_req_implemented.append(row.total_sw_req_implemented)
        ccm_more_50.append(row.ccm_more_50)
        ccm_24_50.append(row.ccm_24_50)
        ccm_12_24.append(row.ccm_12_24)
        misra_high.append(row.misra_high)
        misra_mid.append(row.misra_mid)
        misra_low.append(row.misra_low)

    return jsonify(
        {'cw': cw,
         'total_client_req': total_client_req,
         'total_change_requests': total_change_requests,
         'change_request_not_reviewed': change_request_not_reviewed,
         'traceability_sys_client': traceability_sys_client,
         'traceability_sw_sys': traceability_sw_sys,
         'total_features': total_features,
         'total_bugs': total_bugs,
         'total_bugs_solved': total_bugs_solved,
         'total_problems': total_problems,
         'total_problems_solved': total_problems_solved,
         'traceability_client_sys': traceability_client_sys,
         'traceability_sys_sw': traceability_sys_sw,
         'total_sw_req_tests': total_sw_req_tests,
         'total_sw_req_passed': total_sw_req_passed,
         'total_sys_req_tests': total_sys_req_tests,
         'total_sys_req_tests_passed': total_sys_req_tests_passed,
         'branch_coverage': branch_coverage,
         'line_coverage': line_coverage,
         'mc_dc_coverage': mc_dc_coverage,
         'total_sw_req_approved': total_sw_req_approved,
         'total_client_req_approved': total_client_req_approved,
         'total_open_issue_client_req': total_open_issue_client_req,
         'total_sys_req': total_sys_req,
         'total_sys_req_implemented': total_sys_req_implemented,
         'total_sw_req': total_sw_req,
         'total_sys_req_approved': total_sys_req_approved,
         'total_sw_req_implemented': total_sw_req_implemented,
         'ccm_more_50': ccm_more_50,
         'ccm_24_50': ccm_24_50,
         'ccm_12_24': ccm_12_24,
         'misra_high': misra_high,
         'misra_mid': misra_mid,
         'misra_low': misra_low,
         })


@bp_input.route('/generate')
def generate():
    raw_data = Workbook.query.all()

    filename = os.path.join(os.path.join(os.path.abspath(os.path.dirname(__file__)), '..', 'uploads'), 'excel.xlsx')
    wb = load_workbook(filename)
    ws = wb.active

    columns = list(itertools.chain(string.ascii_uppercase,
                                   (''.join(pair) for pair in itertools.product(string.ascii_uppercase, repeat=2))))

    col = 0
    for item in raw_data:
        col += 1
        ws[f'{columns[col]}1'] = f'id{col}'
        ws[f'{columns[col]}2'] = item.cw
        ws[f'{columns[col]}3'] = item.total_client_req
        ws[f'{columns[col]}4'] = item.total_client_req_approved
        ws[f'{columns[col]}5'] = item.total_open_issue_client_req
        ws[f'{columns[col]}6'] = item.total_sys_req
        ws[f'{columns[col]}7'] = item.total_sys_req_approved
        ws[f'{columns[col]}8'] = item.total_sys_req_implemented
        ws[f'{columns[col]}9'] = item.total_sw_req
        ws[f'{columns[col]}10'] = item.total_sw_req_approved
        ws[f'{columns[col]}11'] = item.total_sw_req_implemented
        ws[f'{columns[col]}12'] = item.ccm_more_50
        ws[f'{columns[col]}13'] = item.ccm_24_50
        ws[f'{columns[col]}14'] = item.ccm_12_24
        ws[f'{columns[col]}15'] = item.misra_high
        ws[f'{columns[col]}16'] = item.misra_mid
        ws[f'{columns[col]}17'] = item.misra_low
        ws[f'{columns[col]}18'] = item.branch_coverage
        ws[f'{columns[col]}19'] = item.line_coverage
        ws[f'{columns[col]}20'] = item.mc_dc_coverage
        ws[f'{columns[col]}21'] = item.total_sw_req_tests
        ws[f'{columns[col]}22'] = item.total_sw_req_passed
        ws[f'{columns[col]}23'] = item.total_sys_req_tests
        ws[f'{columns[col]}24'] = item.total_sys_req_tests_passed
        ws[f'{columns[col]}25'] = item.traceability_sys_client
        ws[f'{columns[col]}26'] = item.traceability_client_sys
        ws[f'{columns[col]}27'] = item.traceability_sys_sw
        ws[f'{columns[col]}28'] = item.traceability_sw_sys
        ws[f'{columns[col]}29'] = item.total_features
        ws[f'{columns[col]}30'] = item.total_bugs
        ws[f'{columns[col]}31'] = item.total_bugs_solved
        ws[f'{columns[col]}32'] = item.total_problems
        ws[f'{columns[col]}33'] = item.total_problems_solved
        ws[f'{columns[col]}34'] = item.total_change_requests
        ws[f'{columns[col]}35'] = item.change_request_not_reviewed
        wb.save(filename)
    return redirect(url_for("input.download"))


@bp_input.route('/download')
def download():
    path = os.path.join(os.path.join(os.path.abspath(os.path.dirname(__file__)), '..', 'uploads'), 'excel.xlsx')
    return send_file(path)


'''
@bp_main.route('/charts', methods=['GET', "POST"])
def charts():
    shrd = []
    data = Workbook.all_of_workbooks()
    for row in data:
        shrd.append([row.cw, row.total_client_req, row.total_client_req_approved])
    shrd.insert(0, ["CW", "Total_Shrd", "Total ShRd Approved"])
    shrd = jsonify(shrd)
    data = {'Task': 'Hours per Day', 'Work': 11, 'Eat': 2, 'Commute': 2, 'Watching TV': 2, 'Sleeping': 7}
    data = [{'Task': 'Hours per Day', 'Work': 11, 'Eat': 2, 'Commute': 2, 'Watching TV': 2, 'Sleeping': 7},
            {'Work': 12, 'Eat': 4, 'Commute': 5, 'Watching TV': 8, 'Sleeping': 9}]

    return render_template('charts_google.html', raw_data=Workbook.all_of_workbooks(), shrd=shrd, datas=data)
'''
